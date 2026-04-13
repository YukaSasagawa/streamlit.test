import openpyxl
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from scipy.optimize import curve_fit
import io
import re

def find_column_by_header(ws, header_row, target_text):
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value == target_text:
            return col
    return None


def find_column_by_header_contains(ws, header_row, target_text):
    target_text = str(target_text).strip()

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue

        value_text = str(value).strip()
        if target_text in value_text:
            return col

    return None

def extract_segment_name_only(full_name):
    """
    例:
    'A企業（M1）' -> 'M1'
    'A企業(F1層)' -> 'F1層'
    """
    if full_name is None:
        return ""

    text = str(full_name).strip()

    m = re.search(r"[（(](.*?)[）)]", text)
    if m:
        return m.group(1).strip()

    return text

def read_tval_reach_data(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 16行目の見出し位置を動的に取得
    grp_start_col = find_column_by_header_contains(ws, 16, "累計GRP")
    reach_count_start_col = find_column_by_header(ws, 16, "エフェクティブリーチ人数（人）")
    reach_rate_start_col = find_column_by_header(ws, 16, "エフェクティブリーチ率（%）")
    cpe_start_col = find_column_by_header(ws, 16, "CPE（円）")

    if None in [grp_start_col, reach_count_start_col, reach_rate_start_col, cpe_start_col]:
        raise ValueError("16行目の見出し（累計GRP/リーチ人数/リーチ率/CPE）の位置を特定できませんでした。")

    # 対象列範囲
    grp_cols = list(range(grp_start_col, reach_count_start_col))
    reach_count_cols = list(range(reach_count_start_col, reach_rate_start_col))
    reach_rate_cols = list(range(reach_rate_start_col, cpe_start_col))

    # データ行範囲: 18行目以降、A列が空になるまで
    start_row = 18
    end_row = start_row
    while ws.cell(row=end_row, column=1).value is not None:
        end_row += 1
    end_row -= 1

    segments = []

    # 3ブロックが同じ順番で並んでいる前提で対応づける
    for grp_col, count_col, rate_col in zip(grp_cols, reach_count_cols, reach_rate_cols):
        segment_name_grp = ws.cell(row=17, column=grp_col).value
        segment_name_count = ws.cell(row=17, column=count_col).value
        segment_name_rate = ws.cell(row=17, column=rate_col).value

        if segment_name_grp is None and segment_name_count is None and segment_name_rate is None:
            continue

        segment_name = (
            segment_name_grp
            if segment_name_grp is not None
            else segment_name_count
            if segment_name_count is not None
            else segment_name_rate
        )
        segment_name = str(segment_name).strip()

        grp_values = []
        count_values = []
        rate_values = []

        for row in range(start_row, end_row + 1):
            grp_val = pd.to_numeric(ws.cell(row=row, column=grp_col).value, errors="coerce")
            count_val = pd.to_numeric(ws.cell(row=row, column=count_col).value, errors="coerce")
            rate_val = pd.to_numeric(ws.cell(row=row, column=rate_col).value, errors="coerce")

            grp_values.append(grp_val)
            count_values.append(count_val)
            rate_values.append(rate_val)

        segment_df = pd.DataFrame({
            "grp": grp_values,
            "reach_count": count_values,
            "reach_rate": rate_values,
        }).dropna(subset=["grp"])

        segments.append({
            "segment": segment_name,
            "segment_name_only": extract_segment_name_only(segment_name),
            "x": segment_df["grp"].tolist(),
            "reach_count": segment_df["reach_count"].tolist(),
            "reach_rate": segment_df["reach_rate"].tolist(),
        })

    return {
        "segments": segments
    }


def hill_function(x, L, K, p, y0):
    x = np.asarray(x, dtype=float)
    return y0 + (L - y0) * ((x ** p) / ((K ** p) + (x ** p)))


def build_hill_saturation_curve(x_values, y_values, num_points=300, y_max=100):
    df = pd.DataFrame({"x": x_values, "y": y_values}).dropna()

    if df.empty:
        return [], []

    df = (
        df.sort_values("x")
          .groupby("x", as_index=False)["y"]
          .mean()
    )

    x = df["x"].to_numpy(dtype=float)
    y = df["y"].to_numpy(dtype=float)

    x = np.maximum(x, 0)

    if y_max is not None:
        y = np.clip(y, 0, y_max)
    else:
        y = np.maximum(y, 0)

    if len(x) == 1:
        return x.tolist(), y.tolist()

    if np.allclose(y, y[0]):
        xs = np.linspace(0, x.max(), num_points)
        ys = np.full_like(xs, max(float(y[0]), 0.0), dtype=float)
        return xs.tolist(), ys.tolist()

    y0 = float(y[0])
    ymax = float(np.max(y))

    positive_x = x[x > 0]
    if len(positive_x) > 0:
        k0 = float(np.median(positive_x))
    else:
        k0 = max(float(np.max(x)) / 2, 1.0)

    p0 = [
        max(ymax + 1.0, y0 + 1.0),  # L
        max(k0, 1e-6),              # K
        1.5,                        # p
    ]

    if y_max is not None:
        l_upper = float(y_max)
    else:
        l_upper = max(ymax * 3.0, y0 + 1.0, 10.0)

    lower_bounds = [
        max(ymax, y0 + 1e-6),  # L
        1e-6,                  # K
        0.3,                   # p
    ]
    upper_bounds = [
        l_upper,                              # L
        max(float(np.max(x)) * 3.0, 10.0),   # K
        6.0,                                 # p
    ]

    try:
        popt, _ = curve_fit(
            lambda xx, L, K, p: hill_function(xx, L, K, p, y0),
            x,
            y,
            p0=p0,
            bounds=(lower_bounds, upper_bounds),
            maxfev=20000,
        )

        L, K, p = popt

        xs = np.linspace(0, x.max(), num_points)
        ys = hill_function(xs, L, K, p, y0)

        if y_max is not None:
            ys = np.clip(ys, 0, y_max)
        else:
            ys = np.maximum(ys, 0)

        return xs.tolist(), ys.tolist()

    except Exception:
        xs = np.linspace(0, x.max(), num_points)
        ys = np.interp(xs, x, y, left=y[0], right=y[-1])

        if y_max is not None:
            ys = np.clip(ys, 0, y_max)
        else:
            ys = np.maximum(ys, 0)

        return xs.tolist(), ys.tolist()


def build_reach_hover_points(
    smooth_x,
    smooth_y,
    grp_tick_step,
    increase_unit="ポイント",
    decimals=2,
    use_comma=False,
):
    if not smooth_x or not smooth_y:
        return [], [], []

    x_max = float(max(smooth_x))

    hover_x = np.arange(0, x_max + 1e-9, grp_tick_step, dtype=float)
    hover_y = np.interp(hover_x, smooth_x, smooth_y)

    rise_points = []
    for i, current_y in enumerate(hover_y):
        if i == 0:
            rise_points.append("—")
        else:
            prev_y = hover_y[i - 1]
            diff = current_y - prev_y

            if use_comma:
                rise_points.append(f"{diff:,.0f}{increase_unit}")
            else:
                rise_points.append(f"{diff:.{decimals}f}{increase_unit}")

    return hover_x.tolist(), hover_y.tolist(), rise_points


def build_reach_line_figure(
    reach_data,
    selected_segment_names=None,
    grp_tick_step=50,
    y_field="reach_rate",
    y_axis_title="リーチ率（%）",
    hover_value_label="リーチ率",
    hover_increase_label="上昇ポイント",
    increase_unit="ポイント",
    y_max=100,
    value_decimals=2,
    value_suffix="%",
    use_comma_for_value=False,
    use_comma_for_increase=False,
    y_tickformat=None,
):
    segments = reach_data["segments"]
    selected_segment_names = set(selected_segment_names or [])

    smooth_palette = [
        "#1f77b4",
        "#ff7f0e",
        "#2ca02c",
        "#d62728",
        "#9467bd",
        "#8c564b",
        "#e377c2",
        "#7f7f7f",
        "#bcbd22",
        "#17becf",
    ]

    segment_color_map = {
        seg["segment"]: smooth_palette[idx % len(smooth_palette)]
        for idx, seg in enumerate(segments)
    }

    fig = go.Figure()
    x_axis_max = 0

    for seg in segments:
        select_label = seg["segment_name_only"]
        legend_label = seg["segment"]

        if select_label not in selected_segment_names:
            continue

        raw_df = pd.DataFrame({
            "grp": seg["x"],
            "y": seg[y_field],
        }).dropna()

        if raw_df.empty:
            continue

        raw_df = (
            raw_df.sort_values("grp")
                  .groupby("grp", as_index=False)["y"]
                  .mean()
        )

        raw_x = raw_df["grp"].tolist()
        raw_y = raw_df["y"].tolist()

        smooth_x, smooth_y = build_hill_saturation_curve(
            raw_x,
            raw_y,
            num_points=300,
            y_max=y_max,
        )

        hover_x, hover_y, rise_points = build_reach_hover_points(
            smooth_x,
            smooth_y,
            grp_tick_step,
            increase_unit=increase_unit,
            decimals=value_decimals,
            use_comma=use_comma_for_increase,
        )

        if smooth_x:
            x_axis_max = max(x_axis_max, max(smooth_x))

        if use_comma_for_value:
            value_template = "%{y:,.0f}"
        else:
            value_template = f"%{{y:.{value_decimals}f}}"

        fig.add_trace(
            go.Scatter(
                x=smooth_x,
                y=smooth_y,
                mode="lines",
                name=legend_label,
                legendgroup=legend_label,
                showlegend=True,
                line=dict(
                    color=segment_color_map[seg["segment"]],
                    width=3,
                ),
                hoverinfo="skip",
            )
        )

        fig.add_trace(
            go.Scatter(
                x=hover_x,
                y=hover_y,
                mode="markers",
                name=legend_label,
                legendgroup=legend_label,
                showlegend=False,
                marker=dict(
                    size=12,
                    color="rgba(0,0,0,0)",
                ),
                customdata=np.array(rise_points, dtype=object),
                hovertemplate=(
                    f"セグメント:{seg['segment']}<br>"
                    "出稿量:%{x:.0f}GRP<br>"
                    f"{hover_value_label}:{value_template}{value_suffix}<br>"
                    f"{hover_increase_label}:%{{customdata}}<extra></extra>"
                ),
            )
        )

    fig.update_layout(
        xaxis_title="累計出稿量",
        yaxis_title=y_axis_title,
        height=650,
        margin=dict(l=50, r=20, t=20, b=120),
        hovermode="closest",
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.18,
            xanchor="center",
            x=0.5,
            title=None,
            traceorder="normal",
        ),
    )

    fig.update_xaxes(
        dtick=grp_tick_step,
        tick0=0,
        range=[0, x_axis_max if x_axis_max > 0 else grp_tick_step],
        automargin=True,
    )

    if y_max is not None:
        fig.update_yaxes(
            automargin=True,
            range=[0, y_max],
            tickformat=y_tickformat,
            separatethousands=True,
        )
    else:
        fig.update_yaxes(
            automargin=True,
            rangemode="tozero",
            tickformat=y_tickformat,
            separatethousands=True,
        )

    return fig