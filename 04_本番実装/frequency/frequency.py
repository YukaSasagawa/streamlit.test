import openpyxl
import pandas as pd
import plotly.graph_objects as go
import io
import re

def extract_segment_name_only(full_name):
    """
    例:
    'A企業（M1）' -> 'M1'
    'A企業(F1層)' -> 'F1層'
    """
    if full_name is None:
        return ""

    text = str(full_name).strip()

    m = re.search(r"[（(](.*?)[）)]", text)
    if m:
        return m.group(1).strip()

    return text

def read_tval_frequency_data(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # A19:A23 の区分
    freq_labels = [ws.cell(row=r, column=1).value for r in range(19, 24)]
    freq_labels = [str(v).strip() if v is not None else "" for v in freq_labels]

    segments = []
    col = 2  # B列スタート

    while True:
        segment_name = ws.cell(row=17, column=col).value
        ratio_values = [ws.cell(row=r, column=col).value for r in range(19, 24)]

        # セグメント名も割合も空なら終了
        if segment_name is None and all(v is None for v in ratio_values):
            break

        # セグメント名がある列だけ採用（B, D, F...）
        if segment_name is not None:
            segment_name = str(segment_name).strip()
            ratios = [pd.to_numeric(v, errors="coerce") for v in ratio_values]

            segments.append({
                "segment": segment_name,
                "segment_name_only": extract_segment_name_only(segment_name),
                "ratios": ratios,
            })

        col += 2  # 割合列を1列飛ばしで見ていく

    return {
        "labels": freq_labels,
        "segments": segments,
    }

def wrap_text_every_n_chars(text, n=10):
    if text is None:
        return ""
    text = str(text)
    return "<br>".join(text[i:i+n] for i in range(0, len(text), n))


def format_frequency_x_label(full_name):
    """
    基本は [企業名] と [（セグメント名）] の間で改行。
    さらに長い場合はそれぞれの塊の中でも改行。
    """
    if full_name is None:
        return ""

    text = str(full_name).strip()

    m = re.match(r"^(.*?)([（(].*[）)])$", text)
    if m:
        company = m.group(1).strip()
        segment = m.group(2).strip()

        company_wrapped = wrap_text_every_n_chars(company, n=10)
        segment_wrapped = wrap_text_every_n_chars(segment, n=10)

        if company_wrapped and segment_wrapped:
            return company_wrapped + "<br>" + segment_wrapped
        return company_wrapped or segment_wrapped

    return wrap_text_every_n_chars(text, n=10)


def build_frequency_stacked_bar_figure(freq_data, selected_segment_names=None):
    labels = freq_data["labels"]
    segments = freq_data["segments"]

    if selected_segment_names:
        segments = [
            s for s in segments
            if s["segment_name_only"] in selected_segment_names
        ]

    color_map = ["#ff8f86", "#ff5050", "#0071bc", "#00468b", "#00215d"]

    x_values = [format_frequency_x_label(s["segment"]) for s in segments]
    raw_segment_names = [s["segment"] for s in segments]

    fig = go.Figure()

    # 棒そのものの幅
    bar_width = 0.55

    for idx, label in enumerate(labels):
        y_values = []
        text_values = []

        for seg in segments:
            value = seg["ratios"][idx]
            value = 0 if pd.isna(value) else float(value)
            y_values.append(value)

            if value >= 4:
                text_values.append(f"{value:.2f}")
            else:
                text_values.append("")

        fig.add_trace(
            go.Bar(
                name=label,
                x=x_values,
                y=y_values,
                marker_color=color_map[idx],
                text=text_values,
                textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="white", size=11),
                customdata=raw_segment_names,
                width=bar_width,
                hovertemplate=(
                    "セグメント:%{customdata}<br>"
                    f"区分:{label}<br>"
                    "割合:%{y:.2f}%<extra></extra>"
                ),
            )
        )

    fig.update_layout(
        barmode="stack",
        xaxis_title=None,
        yaxis_title="割合（%）",
        yaxis=dict(range=[0, 100]),
        height=650,
        margin=dict(l=40, r=20, t=20, b=140),
        bargap=0.15,
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.18,
            xanchor="center",
            x=0.5,
            title=None,
            traceorder="normal",
        ),
    )

    if len(segments) <= 4:
        visible_slots = 5
        side_padding = (visible_slots - len(segments)) / 2

        fig.update_xaxes(
            tickangle=0,
            automargin=True,
            range=[
                -0.5 - side_padding,
                (len(segments) - 0.5) + side_padding,
            ],
        )
    else:
        fig.update_xaxes(
            tickangle=0,
            automargin=True,
        )

    return fig