import io
from datetime import datetime, time, timedelta

import openpyxl
import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
import streamlit as st


st.set_page_config(page_title="VR / TVAL / REVISIO ヒートマップ", layout="wide")
st.title("VR / TVAL / REVISIO ヒートマップ表示")


def format_label(value):
    if value is None:
        return ""

    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        total_hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{total_hours}:{minutes:02d}"

    if isinstance(value, time):
        return f"{value.hour}:{value.minute:02d}"

    if isinstance(value, datetime):
        return f"{value.hour}:{value.minute:02d}"

    if isinstance(value, str):
        return value.strip()

    return str(value)


def extract_block(
    ws,
    display_title,
    header_row,
    header_start_col,
    header_end_col,
    index_col,
    data_start_row,
    data_end_row,
    data_start_col,
    data_end_col,
):
    col_labels = [
        ws.cell(row=header_row, column=col).value
        for col in range(header_start_col, header_end_col + 1)
    ]
    row_labels = [
        ws.cell(row=row, column=index_col).value
        for row in range(data_start_row, data_end_row + 1)
    ]
    values = [
        [ws.cell(row=row, column=col).value for col in range(data_start_col, data_end_col + 1)]
        for row in range(data_start_row, data_end_row + 1)
    ]

    col_labels = [format_label(v) for v in col_labels]
    row_labels = [format_label(v) for v in row_labels]

    df = pd.DataFrame(values, index=row_labels, columns=col_labels)
    df = df.apply(pd.to_numeric, errors="coerce")

    return {
        "title": display_title,
        "data": df,
    }


def read_vr_heatmaps(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    heatmaps = [
        extract_block(ws, "NTV", 6, 2, 8, 1, 7, 30, 2, 8),
        extract_block(ws, "TBS", 6, 28, 34, 27, 7, 30, 28, 34),
        extract_block(ws, "CX", 32, 15, 21, 14, 33, 56, 15, 21),
        extract_block(ws, "EX", 6, 15, 21, 14, 7, 30, 15, 21),
        extract_block(ws, "TX", 32, 2, 8, 1, 33, 56, 2, 8),
    ]
    return heatmaps


def read_tval_heatmaps(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    heatmaps = [
        extract_block(ws, "NTV", 13, 10, 16, 1, 14, 37, 10, 16),
        extract_block(ws, "TBS", 13, 26, 32, 1, 14, 37, 26, 32),
        extract_block(ws, "CX", 13, 42, 48, 1, 14, 37, 42, 48),
        extract_block(ws, "EX", 13, 18, 24, 1, 14, 37, 18, 24),
        extract_block(ws, "TX", 13, 34, 40, 1, 14, 37, 34, 40),
    ]
    return heatmaps


def read_revisio_heatmaps(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    sheet_order = ["NTV", "TBS", "CX", "EX", "TX"]
    heatmaps = []

    for sheet_name in sheet_order:
        ws = wb[sheet_name]
        heatmaps.append(
            extract_block(ws, sheet_name, 1, 3, 9, 2, 2, 25, 3, 9)
        )

    return heatmaps


def get_rank_series(df, decimals=2):
    values = df.stack()

    if values.empty:
        return values

    return values.round(decimals)


def build_wish_frame(df, worst_n):
    raw_values = df.stack()
    rank_values = get_rank_series(df, decimals=2)

    if raw_values.empty:
        return pd.DataFrame("", index=df.index, columns=df.columns)

    mean_val = rank_values.mean()
    worst_n = max(0, min(int(worst_n), len(rank_values)))

    worst_positions = set()
    if worst_n > 0:
        sorted_values = rank_values.sort_values(ascending=True)
        threshold = sorted_values.iloc[worst_n - 1]
        worst_positions = set(rank_values[rank_values <= threshold].index.tolist())

    result = pd.DataFrame("", index=df.index, columns=df.columns)

    for row_label in df.index:
        for col_label in df.columns:
            value = df.loc[row_label, col_label]
            rank_value = round(value, 2) if pd.notna(value) else None

            if pd.isna(value):
                result.loc[row_label, col_label] = ""
            elif (row_label, col_label) in worst_positions:
                result.loc[row_label, col_label] = "×"
            elif rank_value > mean_val:
                result.loc[row_label, col_label] = "〇"
            else:
                result.loc[row_label, col_label] = ""

    return result


def dataframe_to_csv_bytes(df):
    return df.to_csv(index=True, encoding="utf-8-sig").encode("utf-8-sig")


def build_wish_table_figure(wish_df, title):
    header_values = [""] + list(wish_df.columns)
    cell_values = [list(wish_df.index)] + [wish_df[col].fillna("").tolist() for col in wish_df.columns]

    n_rows = len(wish_df)
    table_height = max(680, 90 + n_rows * 27)

    fig = go.Figure(
        data=[
            go.Table(
                columnwidth=[70] + [42] * len(wish_df.columns),
                header=dict(
                    values=header_values,
                    align="center",
                    height=30,
                ),
                cells=dict(
                    values=cell_values,
                    align="center",
                    height=27,
                ),
            )
        ]
    )

    fig.update_layout(
        title=title,
        title_x=0.5,
        height=table_height,
        margin=dict(l=5, r=5, t=45, b=5),
        font=dict(size=11),
    )

    return fig


def figure_to_png_bytes(fig):
    try:
        return pio.to_image(fig, format="png", scale=2)
    except Exception:
        return None


def render_wish_row(section_title, heatmaps, worst_n):
    st.markdown(f"### {section_title}")
    cols = st.columns(5)

    for idx, (col, hm) in enumerate(zip(cols, heatmaps)):
        with col:
            wish_df = build_wish_frame(hm["data"], worst_n)
            wish_fig = build_wish_table_figure(wish_df, hm["title"])

            png_bytes = figure_to_png_bytes(wish_fig)
            csv_bytes = dataframe_to_csv_bytes(wish_df)

            btn_col1, btn_col2 = st.columns(2)
            with btn_col1:
                if png_bytes is not None:
                    st.download_button(
                        label="PNG",
                        data=png_bytes,
                        file_name=f"{section_title}_{hm['title']}_wish.png",
                        mime="image/png",
                        key=f"wish_png_{section_title}_{hm['title']}_{idx}",
                        use_container_width=True,
                        on_click="ignore",
                    )
                else:
                    st.button(
                        "PNG",
                        disabled=True,
                        key=f"wish_png_disabled_{section_title}_{hm['title']}_{idx}",
                        use_container_width=True,
                    )

            with btn_col2:
                st.download_button(
                    label="CSV",
                    data=csv_bytes,
                    file_name=f"{section_title}_{hm['title']}_wish.csv",
                    mime="text/csv",
                    key=f"wish_csv_{section_title}_{hm['title']}_{idx}",
                    use_container_width=True,
                    on_click="ignore",
                )

            st.plotly_chart(
                wish_fig,
                use_container_width=True,
                config={
                    "displayModeBar": False,
                    "displaylogo": False,
                },
                key=f"wish_table_{section_title}_{hm['title']}_{idx}",
            )

    if any(figure_to_png_bytes(build_wish_table_figure(build_wish_frame(hm["data"], worst_n), hm["title"])) is None for hm in heatmaps):
        st.caption("PNGダウンロードには kaleido が必要です")


def get_rank_positions(df, top_n, worst_n):
    rank_values = get_rank_series(df, decimals=2)

    if rank_values.empty:
        return set(), set()

    top_n = max(0, min(int(top_n), len(rank_values)))
    worst_n = max(0, min(int(worst_n), len(rank_values)))

    top_positions = set()
    worst_positions = set()

    if top_n > 0:
        sorted_desc = rank_values.sort_values(ascending=False)
        top_threshold = sorted_desc.iloc[top_n - 1]
        top_positions = set(rank_values[rank_values >= top_threshold].index.tolist())

    if worst_n > 0:
        sorted_asc = rank_values.sort_values(ascending=True)
        worst_threshold = sorted_asc.iloc[worst_n - 1]
        worst_positions = set(rank_values[rank_values <= worst_threshold].index.tolist())

    return top_positions, worst_positions


def build_heatmap_figure(hm, top_n=0, worst_n=0):
    df = hm["data"]
    nrows, ncols = df.shape

    customdata = [
        [[str(df.columns[c]), str(df.index[r])] for c in range(ncols)]
        for r in range(nrows)
    ]

    text_values = [
        [("" if pd.isna(v) else f"{v:.2f}") for v in row]
        for row in df.values
    ]

    fig = go.Figure(
        data=go.Heatmap(
            z=df.values,
            x=list(range(ncols)),
            y=list(range(nrows)),
            colorscale="Blues",
            showscale=False,
            text=text_values,
            texttemplate="%{text}",
            textfont={"size": 8},
            customdata=customdata,
            hovertemplate=(
                f"放送局:{hm['title']}"
                f"<br>曜日:%{{customdata[0]}}"
                f"<br>時間帯:%{{customdata[1]}}"
                f"<br>値:%{{z:.2f}}"
                f"<extra></extra>"
            ),
        )
    )

    fig.update_layout(
        height=420,
        title=str(hm["title"]),
        title_x=0.5,
        margin=dict(l=5, r=5, t=60, b=5),
        font=dict(size=10),
    )

    fig.update_xaxes(
        tickmode="array",
        tickvals=list(range(ncols)),
        ticktext=list(df.columns),
        side="top",
        tickfont=dict(size=9),
        title=None,
    )

    fig.update_yaxes(
        tickmode="array",
        tickvals=list(range(nrows)),
        ticktext=list(df.index),
        tickfont=dict(size=9),
        title=None,
        autorange="reversed",
    )

    top_positions, worst_positions = get_rank_positions(df, top_n, worst_n)

    row_to_idx = {label: i for i, label in enumerate(df.index)}
    col_to_idx = {label: i for i, label in enumerate(df.columns)}

    for row_label, col_label in top_positions:
        r = row_to_idx[row_label]
        c = col_to_idx[col_label]

        fig.add_shape(
            type="rect",
            x0=c - 0.5,
            x1=c + 0.5,
            y0=r - 0.5,
            y1=r + 0.5,
            line=dict(color="red", width=3),
            fillcolor="rgba(0,0,0,0)",
            layer="above",
        )

    for row_label, col_label in worst_positions:
        r = row_to_idx[row_label]
        c = col_to_idx[col_label]

        fig.add_shape(
            type="rect",
            x0=c - 0.5,
            x1=c + 0.5,
            y0=r - 0.5,
            y1=r + 0.5,
            line=dict(color="blue", width=3),
            fillcolor="rgba(0,0,0,0)",
            layer="above",
        )

    return fig


def render_heatmap_row(section_title, heatmaps, station_filters):
    st.markdown(f"### {section_title}")
    cols = st.columns(5)

    for idx, (col, hm) in enumerate(zip(cols, heatmaps)):
        with col:
            station = hm["title"]
            top_n = station_filters[station]["top"]
            worst_n = station_filters[station]["worst"]

            fig = build_heatmap_figure(hm, top_n=top_n, worst_n=worst_n)

            st.plotly_chart(
                fig,
                use_container_width=True,
                config={
                    "displayModeBar": True,
                    "displaylogo": False,
                },
                key=f"{section_title}_{hm['title']}_{idx}",
            )


tab_upload, tab_heatmap, tab_wish = st.tabs(["アップロード", "ヒートマップ", "希望枠"])


with tab_upload:
    st.subheader("Excelファイルをアップロード")

    st.markdown("#### VRデータ")
    st.file_uploader(
        "VRのExcelファイルを選択してください",
        type=["xlsx", "xlsm"],
        key="vr_file",
    )
    if st.session_state.get("vr_file") is not None:
        st.success(f"アップロード完了: {st.session_state['vr_file'].name}")

    st.markdown("#### TVALデータ")
    st.file_uploader(
        "TVALのExcelファイルを選択してください",
        type=["xlsx", "xlsm"],
        key="tval_file",
    )
    if st.session_state.get("tval_file") is not None:
        st.success(f"アップロード完了: {st.session_state['tval_file'].name}")

    st.markdown("#### REVISIOデータ")
    st.file_uploader(
        "REVISIOのExcelファイルを選択してください",
        type=["xlsx", "xlsm"],
        key="revisio_file",
    )
    if st.session_state.get("revisio_file") is not None:
        st.success(f"アップロード完了: {st.session_state['revisio_file'].name}")


with tab_heatmap:
    st.subheader("ヒートマップ")

    st.markdown("#### 放送局別フィルタ")
    filter_cols = st.columns(5)
    station_filters = {}

    for col, station in zip(filter_cols, ["NTV", "TBS", "CX", "EX", "TX"]):
        with col:
            st.markdown(f"**{station}**")
            top_n = st.number_input(
                f"{station} TOP",
                min_value=0,
                max_value=50,
                value=0,
                step=1,
                key=f"top_{station}",
            )
            worst_n = st.number_input(
                f"{station} WORST",
                min_value=0,
                max_value=50,
                value=0,
                step=1,
                key=f"worst_{station}",
            )

            station_filters[station] = {
                "top": top_n,
                "worst": worst_n,
            }

    st.caption("赤枠: TOP指定件数 / 青枠: WORST指定件数")
    st.caption("各ヒートマップ右上のカメラアイコンからPNG保存できます")

    vr_file = st.session_state.get("vr_file")
    tval_file = st.session_state.get("tval_file")
    revisio_file = st.session_state.get("revisio_file")

    if vr_file is None and tval_file is None and revisio_file is None:
        st.info("先に『アップロード』タブでVRデータ、TVALデータ、REVISIOデータをアップロードしてください。")
    else:
        if vr_file is not None:
            try:
                vr_heatmaps = read_vr_heatmaps(vr_file)
                render_heatmap_row("VR", vr_heatmaps, station_filters)
            except Exception as e:
                st.error("VRデータの読み取りに失敗しました。")
                st.exception(e)
        else:
            st.warning("VRデータが未アップロードです。")

        st.markdown("---")

        if tval_file is not None:
            try:
                tval_heatmaps = read_tval_heatmaps(tval_file)
                render_heatmap_row("TVAL", tval_heatmaps, station_filters)
            except Exception as e:
                st.error("TVALデータの読み取りに失敗しました。")
                st.exception(e)
        else:
            st.warning("TVALデータが未アップロードです。")

        st.markdown("---")

        if revisio_file is not None:
            try:
                revisio_heatmaps = read_revisio_heatmaps(revisio_file)
                render_heatmap_row("REVISIO", revisio_heatmaps, station_filters)
            except Exception as e:
                st.error("REVISIOデータの読み取りに失敗しました。")
                st.exception(e)
        else:
            st.warning("REVISIOデータが未アップロードです。")


with tab_wish:
    st.subheader("希望枠")

    worst_n = st.number_input(
        "ワースト下位の件数",
        min_value=0,
        max_value=50,
        value=5,
        step=1,
    )

    vr_file = st.session_state.get("vr_file")
    tval_file = st.session_state.get("tval_file")
    revisio_file = st.session_state.get("revisio_file")

    if vr_file is None and tval_file is None and revisio_file is None:
        st.info("先に『アップロード』タブでVRデータ、TVALデータ、REVISIOデータをアップロードしてください。")
    else:
        if vr_file is not None:
            try:
                vr_heatmaps = read_vr_heatmaps(vr_file)
                render_wish_row("VR", vr_heatmaps, worst_n)
            except Exception as e:
                st.error("VRデータの希望枠表示に失敗しました。")
                st.exception(e)
        else:
            st.warning("VRデータが未アップロードです。")

        st.markdown("---")

        if tval_file is not None:
            try:
                tval_heatmaps = read_tval_heatmaps(tval_file)
                render_wish_row("TVAL", tval_heatmaps, worst_n)
            except Exception as e:
                st.error("TVALデータの希望枠表示に失敗しました。")
                st.exception(e)
        else:
            st.warning("TVALデータが未アップロードです。")

        st.markdown("---")

        if revisio_file is not None:
            try:
                revisio_heatmaps = read_revisio_heatmaps(revisio_file)
                render_wish_row("REVISIO", revisio_heatmaps, worst_n)
            except Exception as e:
                st.error("REVISIOデータの希望枠表示に失敗しました。")
                st.exception(e)
        else:
            st.warning("REVISIOデータが未アップロードです。")